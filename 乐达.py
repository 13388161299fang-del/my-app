import pandas as pd
from datetime import datetime
import os
import sys
import re
from openpyxl.styles import Alignment, PatternFill, Border, Side

# ================== 配置 ==================
INPUT_DIR = "乐达-表格原文件"          # 存放原始Excel的文件夹
OUTPUT_DIR = "乐达-整理好名单"      # 输出文件夹（自动创建）

# 自动识别列：如果设为 True，则忽略下面的列名变量，自动识别；如果设为 False，则使用下面指定的列名
AUTO_DETECT = True

# 当 AUTO_DETECT = False 时，请根据实际列名修改以下变量
NAME_COL = "姓名"                  # 存放姓名的列名
PHONE_COL = "手机"                  # 存放手机号的列名
ID_COL = "证件号"                    # 存放身份证号的列名
# =========================================

def clean_phone(phone_str):
    """从字符串中提取11位手机号（以1开头）"""
    if not phone_str or pd.isna(phone_str):
        return ''
    s = str(phone_str).strip()
    match = re.search(r'1[3-9]\d{9}', s)
    return match.group() if match else ''

def clean_id(id_str):
    """提取18位身份证号（最后一位可能是数字或X）"""
    if not id_str or pd.isna(id_str):
        return ''
    s = str(id_str).strip().upper()
    match = re.search(r'[1-9]\d{5}(19|20)\d{2}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])\d{3}[\dX]', s)
    return match.group() if match else ''

def calculate_age(id_card):
    """根据18位身份证号计算周岁年龄"""
    if not id_card:
        return None
    try:
        birth_str = id_card[6:14]
        birth_date = datetime.strptime(birth_str, '%Y%m%d')
        today = datetime.now()
        age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        return age
    except:
        return None

def age_category(age):
    """按年龄分类（新规则）"""
    if age is None:
        return '未知'
    if age < 7:
        return '0-7岁'
    elif age < 23:
        return '7-23岁'
    elif age < 60:
        return '23-60岁'
    else:
        return '60岁及以上'

def detect_columns(df):
    """
    自动识别姓名、手机、身份证号所在的列
    返回 (name_col, phone_col, id_col)
    """
    # 对每一列进行评分
    scores = {col: {'name': 0, 'phone': 0, 'id': 0} for col in df.columns}

    for col in df.columns:
        # 取该列所有非空值的前100行作为样本（避免全表扫描太慢）
        sample = df[col].dropna().astype(str).head(100)
        if sample.empty:
            continue

        for value in sample:
            # 手机号评分：包含11位手机号
            if re.search(r'1[3-9]\d{9}', value):
                scores[col]['phone'] += 1
            # 身份证号评分：包含18位身份证号
            if re.search(r'[1-9]\d{5}(19|20)\d{2}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])\d{3}[\dXx]', value):
                scores[col]['id'] += 1
            # 姓名评分：纯2-4个汉字（不考虑其他字符）
            if re.fullmatch(r'[\u4e00-\u9fa5]{2,4}', value):
                scores[col]['name'] += 1

    # 归一化评分（除以样本数，得到比例）
    for col in df.columns:
        total = len(df[col].dropna().head(100))
        if total > 0:
            scores[col]['name'] /= total
            scores[col]['phone'] /= total
            scores[col]['id'] /= total

    # 找出每个类别得分最高的列（要求得分 > 0.3，避免误判）
    name_col = max(scores, key=lambda c: scores[c]['name']) if max(s['name'] for s in scores.values()) > 0.3 else None
    phone_col = max(scores, key=lambda c: scores[c]['phone']) if max(s['phone'] for s in scores.values()) > 0.3 else None
    id_col = max(scores, key=lambda c: scores[c]['id']) if max(s['id'] for s in scores.values()) > 0.3 else None

    # 处理可能同一列被识别为多个类型的情况（优先保留身份证，其次手机，最后姓名）
    if id_col and (id_col == phone_col or id_col == name_col):
        # 如果身份证列与其他冲突，则降低其他列的得分再选
        pass  # 简单处理：如果冲突，我们信任身份证列，其他列从剩余列中选

    # 确保三列不同
    used = set()
    if id_col and id_col not in used:
        used.add(id_col)
    if phone_col and phone_col not in used:
        used.add(phone_col)
    else:
        # 重新选择phone_col（从不在used中的列选最高分）
        candidates = {c: scores[c]['phone'] for c in df.columns if c not in used}
        if candidates:
            phone_col = max(candidates, key=candidates.get)
            if candidates[phone_col] <= 0.3:
                phone_col = None
    if name_col and name_col not in used:
        used.add(name_col)
    else:
        candidates = {c: scores[c]['name'] for c in df.columns if c not in used}
        if candidates:
            name_col = max(candidates, key=candidates.get)
            if candidates[name_col] <= 0.3:
                name_col = None

    # 如果仍有缺失，报错
    missing = []
    if not name_col:
        missing.append("姓名列")
    if not phone_col:
        missing.append("手机列")
    if not id_col:
        missing.append("身份证列")
    if missing:
        raise ValueError(f"自动识别失败，无法确定以下列：{missing}。请检查数据格式或手动设置列名。")

    print(f"自动识别结果：姓名列 = '{name_col}'，手机列 = '{phone_col}'，身份证列 = '{id_col}'")
    return name_col, phone_col, id_col

def process_excel(input_path, output_path):
    print(f"正在处理: {input_path}")
    df = pd.read_excel(input_path)

    # 确定列名
    if AUTO_DETECT:
        name_col, phone_col, id_col = detect_columns(df)
    else:
        name_col, phone_col, id_col = NAME_COL, PHONE_COL, ID_COL
        # 检查手动指定的列是否存在
        missing = []
        if name_col not in df.columns:
            missing.append(name_col)
        if phone_col not in df.columns:
            missing.append(phone_col)
        if id_col not in df.columns:
            missing.append(id_col)
        if missing:
            raise ValueError(f"手动指定的列不存在：{missing}。请修改脚本开头的列名变量或启用自动识别。")

    # 提取并清理数据
    names = df[name_col].astype(str).replace('nan', '').str.strip()
    raw_phones = df[phone_col].astype(str).replace('nan', '')
    raw_ids = df[id_col].astype(str).replace('nan', '')

    phones = [clean_phone(p) for p in raw_phones]
    id_cards = [clean_id(i) for i in raw_ids]

    # 计算年龄和分类
    ages = []
    categories = []
    for id_card in id_cards:
        age = calculate_age(id_card)
        ages.append(age)
        categories.append(age_category(age))

    # 构建基础DataFrame
    base_df = pd.DataFrame({
        '姓名': names,
        '手机': phones,
        '身份证号': id_cards,
        '年龄': ages,
        '年龄段': categories
    })

    # 构建输出DataFrame的各个部分
    parts = []

    # 定义年龄段顺序（与图片代码一致）
    category_order = ['23-60岁', '7-23岁', '0-7岁', '60岁及以上', '未知']

    for i, cat in enumerate(category_order):
        cat_data = base_df[base_df['年龄段'] == cat].copy()
        if cat_data.empty:
            continue

        # 根据年龄段定义细分条件
        sub_condition = pd.Series(False, index=cat_data.index)
        if cat == '23-60岁':
            cond = (cat_data['年龄'] >= 23) & (cat_data['年龄'] < 25)
            sub_condition[cond] = True
        elif cat == '7-23岁':
            cond = (cat_data['年龄'] >= 21) & (cat_data['年龄'] < 23)
            sub_condition[cond] = True

        # 分离主数据和细分数据
        main_data = cat_data[~sub_condition].copy()
        sub_data = cat_data[sub_condition].copy()

        if not main_data.empty:
            parts.append(main_data)
        if not sub_data.empty:
            parts.append(sub_data)

        # 统计行
        total_count = len(cat_data)
        stats_row = pd.DataFrame([{
            '姓名': f'人数：{total_count}',
            '手机': '',
            '身份证号': '',
            '年龄': '',
            '年龄段': ''
        }])
        parts.append(stats_row)

        # 如果不是最后一个年龄段，添加空行
        if i < len(category_order) - 1:
            empty_row = pd.DataFrame([{
                '姓名': '',
                '手机': '',
                '身份证号': '',
                '年龄': '',
                '年龄段': ''
            }])
            parts.append(empty_row)

    # 合并所有部分
    final_df = pd.concat(parts, ignore_index=True)

    # ---------- 写入Excel，设置格式 ----------
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        final_df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # 设置列宽
        col_widths = {'A': 10, 'B': 15, 'C': 30, 'D': 10, 'E': 12}
        for col, width in col_widths.items():
            worksheet.column_dimensions[col].width = width

        # 定义细边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 设置所有单元格居中对齐和边框
        alignment = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = alignment
                cell.border = thin_border

        # 设置手机和身份证号列为文本格式
        for col_idx in [2, 3]:  # B列和C列
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row, col_idx)
                cell.number_format = '@'

        # 定义背景色
        green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # 应用背景色
        for row_idx in range(2, worksheet.max_row + 1):
            name_cell = worksheet.cell(row_idx, 1)
            age_cell = worksheet.cell(row_idx, 4)
            cat_cell = worksheet.cell(row_idx, 5)

            name = name_cell.value
            age = age_cell.value
            cat = cat_cell.value

            if name and str(name).strip() and cat:
                # 0-7岁和60岁及以上为黄色
                if cat == '60岁及以上' or cat == '0-7岁':
                    for col_idx in range(1, 6):
                        worksheet.cell(row_idx, col_idx).fill = yellow_fill
                elif cat == '23-60岁' and age is not None:
                    try:
                        age_val = float(age) if not isinstance(age, int) else age
                        if 23 <= age_val < 25:
                            for col_idx in range(1, 6):
                                worksheet.cell(row_idx, col_idx).fill = green_fill
                    except:
                        pass
                elif cat == '7-23岁' and age is not None:
                    try:
                        age_val = float(age) if not isinstance(age, int) else age
                        if 21 <= age_val < 23:
                            for col_idx in range(1, 6):
                                worksheet.cell(row_idx, col_idx).fill = blue_fill
                    except:
                        pass

    print(f"已完成: {output_path}")

def main():
    if not os.path.isdir(INPUT_DIR):
        print(f"错误：输入目录 '{INPUT_DIR}' 不存在，请检查路径。")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    excel_files = [f for f in os.listdir(INPUT_DIR) if f.lower().endswith('.xlsx')]
    if not excel_files:
        print(f"输入目录 '{INPUT_DIR}' 中没有找到 Excel 文件（.xlsx）。")
        return

    for filename in excel_files:
        input_path = os.path.join(INPUT_DIR, filename)
        output_path = os.path.join(OUTPUT_DIR, filename)

        if os.path.exists(output_path):
            print(f"跳过已处理文件: {filename}")
            continue

        try:
            process_excel(input_path, output_path)
        except Exception as e:
            print(f"处理文件 {filename} 时出错: {e}")

    print("全部处理完毕！")

if __name__ == "__main__":
    main()